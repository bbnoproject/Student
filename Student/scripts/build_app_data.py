from __future__ import annotations

import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "Data"
APP_DIR = ROOT / "App"
OUTPUT_FILE = APP_DIR / "data.js"

EXPECTED_PROJECT_PHASE_LABELS = [
    "1차 프로젝트",
    "2차 프로젝트",
    "3차 프로젝트",
    "4차 프로젝트",
]


PROJECT_PHASES = [
    {"phase": "1차 프로젝트", "daily_index": 0, "retro_index": 1},
    {"phase": "2차 프로젝트", "daily_index": 2, "retro_index": 3},
    {"phase": "3차 프로젝트", "daily_index": 4, "retro_index": None},
]


POSITIVE_COLLAB_KEYWORDS = [
    "소통",
    "협업",
    "조율",
    "도움",
    "배려",
    "공유",
    "책임감",
    "정리",
    "존중",
    "칭찬",
    "조율",
    "함께",
]

NEGATIVE_COLLAB_KEYWORDS = [
    "갈등",
    "충돌",
    "어려움",
    "답답",
    "소음",
    "문제",
    "힘들",
    "불만",
    "재촉",
    "미안",
]

RESILIENCE_KEYWORDS = [
    "극복",
    "해결",
    "다시",
    "개선",
    "보완",
    "화이팅",
    "해보",
    "배웠",
    "성장",
]

LOW_RESILIENCE_KEYWORDS = [
    "포기",
    "무섭",
    "불안",
    "압박",
    "걱정",
    "회피",
    "힘들",
    "어렵",
    "버겁",
]

REFLECTION_KEYWORDS = [
    "느꼈",
    "배웠",
    "다음",
    "앞으로",
    "개선",
    "깨달",
    "생각",
    "반성",
    "보완",
]

CAREER_KEYWORDS = [
    "취업",
    "직무",
    "기획자",
    "포트폴리오",
    "자기소개서",
    "회사",
    "입사",
    "목표",
    "진로",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def compact_text(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value))


def short_text(value: Any, limit: int = 140) -> str:
    text = compact_text(value)
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "…"


def slugify_name(name: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z가-힣]+", "-", name.strip())
    return slug.strip("-").lower()


def normalize_name(name: Any) -> str:
    return re.sub(r"\s+", "", str(name or "")).strip()


def phase_order_key(label: str) -> int:
    match = re.search(r"(\d+)", clean_text(label))
    return int(match.group(1)) if match else 999


def date_in_window(value: date | None, start: date, end: date) -> bool:
    return bool(value and start <= value <= end)


def normalize_project_date(
    raw_value: Any,
    submitted_at: datetime | None,
    course_start: date,
    course_end: date,
) -> date | None:
    parsed = parse_date(raw_value)
    min_date = course_start - timedelta(days=45)
    max_date = course_end + timedelta(days=30)
    timestamp_date = submitted_at.date() if submitted_at else None

    if isinstance(raw_value, str):
        day_match = re.search(r"(\d{1,2})\s*일", clean_text(raw_value))
        if day_match and timestamp_date:
            candidate = date(timestamp_date.year, timestamp_date.month, int(day_match.group(1)))
            if date_in_window(candidate, min_date, max_date):
                return candidate

    if date_in_window(parsed, min_date, max_date):
        return parsed
    if date_in_window(timestamp_date, min_date, max_date):
        return timestamp_date
    if parsed and timestamp_date and parsed.month == timestamp_date.month and parsed.day == timestamp_date.day:
        return timestamp_date
    return parsed


def profile_average(scores: dict[str, int]) -> float:
    return round(mean(scores.values()), 2) if scores else 0.0


def percentile_value(values: list[float], percentile: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    index = max(0, min(len(ordered) - 1, round((len(ordered) - 1) * percentile)))
    return round(ordered[index], 2)


def dominant_phase_dates(dates: list[date]) -> list[date]:
    ordered = sorted(set(dates))
    if len(ordered) <= 1:
      return ordered

    groups: list[list[date]] = [[ordered[0]]]
    for item in ordered[1:]:
        if (item - groups[-1][-1]).days <= 14:
            groups[-1].append(item)
        else:
            groups.append([item])
    groups.sort(key=lambda group: (len(group), group[-1]), reverse=True)
    return sorted(groups[0]) if groups else ordered


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            excel_date = from_excel(value)
            if isinstance(excel_date, datetime):
                return excel_date.date()
            if isinstance(excel_date, date):
                return excel_date
        except Exception:
            pass
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    match = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if match:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return None


def parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time(0, 0))
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y.%m.%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    parsed = parse_date(text)
    if parsed:
        return datetime.combine(parsed, time(0, 0))
    return None


def severity_rank(level: str) -> int:
    return {"info": 1, "stable": 1, "success": 1, "caution": 2, "warning": 3}.get(level, 1)


def event_severity_from_attendance(value: str) -> str:
    if "연락 X" in value or "무단" in value:
        return "warning"
    if "결석" in value:
        return "caution"
    if "지각" in value or "조퇴" in value:
        return "caution"
    return "info"


def attendance_kind(value: str) -> str:
    for key in ["결석", "지각", "조퇴", "외출", "공가", "휴가"]:
        if key in value:
            return key
    return "기타"


def attendance_reason_profile(value: str, kind: str) -> dict[str, str]:
    detail = clean_text(value)
    lowered = detail.lower()
    if not detail or detail == "-":
        return {"category": "unknown", "impact": "contextual", "severity": "info"}

    health_keywords = [
        "아파",
        "병원",
        "몸살",
        "감기",
        "치과",
        "진료",
        "컨디션",
        "병결",
        "검진",
        "입원",
        "치료",
        "간병",
        "병가",
        "건강",
    ]
    administrative_keywords = [
        "국취제",
        "국민취업",
        "실업급여",
        "상담",
        "면접",
        "공가",
        "공결",
        "예비군",
        "민방위",
        "qr",
        "행정",
        "휴학",
    ]
    behavioral_keywords = [
        "연락 x",
        "연락x",
        "연락 안",
        "연락안",
        "연락 안됨",
        "연락안됨",
        "무단",
        "늦잠",
        "피곤",
        "기상 실패",
        "이탈",
    ]
    learning_keywords = ["학업", "시험", "기말고사", "게임잼", "해커톤", "프로젝트", "출장"]
    personal_keywords = [
        "개인 일정",
        "개인 사정",
        "가족 행사",
        "가족 여행",
        "가정사",
        "이사",
        "병문안",
        "가족 모임",
        "본가",
        "외가",
        "행사 방문",
        "여행",
        "휴식",
        "밭일",
    ]
    logistics_keywords = ["버스", "우체국", "지방 출장"]

    if any(keyword in lowered for keyword in health_keywords):
        return {"category": "health", "impact": "contextual", "severity": "info"}
    if any(keyword in lowered for keyword in administrative_keywords):
        return {"category": "administrative", "impact": "exempt", "severity": "info"}
    if any(keyword in lowered for keyword in behavioral_keywords):
        severity = "warning" if kind == "결석" else "caution"
        return {"category": "behavioral", "impact": "behavioral", "severity": severity}
    if any(keyword in lowered for keyword in learning_keywords):
        return {"category": "external_learning", "impact": "contextual", "severity": "info"}
    if any(keyword in lowered for keyword in personal_keywords):
        return {"category": "personal", "impact": "contextual", "severity": "info"}
    if any(keyword in lowered for keyword in logistics_keywords):
        return {"category": "logistics", "impact": "contextual", "severity": "info"}
    if kind == "결석":
        return {"category": "absence", "impact": "contextual", "severity": "caution"}
    if kind in {"지각", "조퇴"}:
        return {"category": "late", "impact": "contextual", "severity": "info"}
    return {"category": "other", "impact": "contextual", "severity": event_severity_from_attendance(detail)}


def score_level(value: int) -> str:
    if value <= 1:
        return "지원 시급"
    if value == 2:
        return "형성 중"
    if value == 3:
        return "안정적"
    return "확장적"


def clamp_score(value: float) -> int:
    return max(1, min(4, int(round(value))))


def keyword_hits(text: str, keywords: list[str]) -> int:
    return sum(text.count(keyword) for keyword in keywords)


def find_header_indexes(headers: list[Any], mapping: dict[str, list[str]]) -> dict[str, int]:
    resolved: dict[str, int] = {}
    cleaned = [compact_text(header) for header in headers]
    for field, keywords in mapping.items():
        resolved[field] = -1
        for idx, header in enumerate(cleaned):
            if any(keyword in header for keyword in keywords):
                resolved[field] = idx
                break
    return resolved


def workbook_by_index(index: int):
    files = sorted([path for path in DATA_DIR.iterdir() if path.suffix.lower() == ".xlsx"])
    return load_workbook(files[index], read_only=True, data_only=True)


@dataclass
class CurriculumWeek:
    week_index: int
    start_date: date
    end_date: date
    label: str
    sessions: list[dict[str, Any]]
    dominant_subject: str
    highlights: list[str]


def build_students() -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    wb = workbook_by_index(8)
    ws = wb.worksheets[0]
    students: list[dict[str, Any]] = []
    by_name: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean_text(row[0])
        phone = clean_text(row[1])
        if not name:
            continue
        student = {
            "id": slugify_name(name),
            "name": name,
            "nameKey": normalize_name(name),
            "phone": phone,
            "gender": clean_text(row[2]),
            "birthDate": parse_date(row[3]).isoformat() if parse_date(row[3]) else "",
            "education": clean_text(row[4]),
            "experience": clean_text(row[5]),
            "address": clean_text(row[6]),
            "specialNote": clean_text(row[7]),
            "attendanceMemo": clean_text(row[8]),
            "cohort": "기획4기",
            "course": "게임 기획 부트캠프",
            "admission": {},
            "timelineEvents": [],
            "checkins": [],
            "retrospectives": [],
            "counselings": [],
            "attendanceEvents": [],
            "projectTeamHistory": [],
            "peerRelationships": [],
            "careerDocuments": {"rounds": [], "summary": {}},
            "evaluationSnapshots": [],
            "statusPeriods": [],
        }
        students.append(student)
        by_name[student["nameKey"]] = student
    return students, by_name


def build_curriculum() -> tuple[dict[str, Any], list[CurriculumWeek]]:
    wb = workbook_by_index(6)
    ws = wb.worksheets[1]
    start = None
    end = None
    sessions: list[dict[str, Any]] = []
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = list(row)
        if row_index == 2:
            start = parse_date(values[2])
        elif row_index == 3:
            end = parse_date(values[2])
        elif row_index >= 6:
            session_date = parse_date(values[2])
            subject = clean_text(values[4])
            title = clean_text(values[7])
            if not session_date or (not subject and not title):
                continue
            sessions.append(
                {
                    "date": session_date.isoformat(),
                    "subject": subject,
                    "lessonTitle": title,
                }
            )

    if not start or not end:
        raise RuntimeError("수업 시간표에서 개강일/종강일을 찾지 못했습니다.")

    weeks: list[CurriculumWeek] = []
    week_index = 1
    cursor = start
    while cursor <= end:
        week_end = min(cursor + timedelta(days=6), end)
        week_sessions = [
            session
            for session in sessions
            if cursor <= parse_date(session["date"]) <= week_end
        ]
        subjects = [session["subject"] for session in week_sessions if session["subject"]]
        dominant_subject = Counter(subjects).most_common(1)[0][0] if subjects else "운영 주간"
        highlight_candidates = []
        seen = set()
        for session in week_sessions:
            lesson_title = session["lessonTitle"]
            if lesson_title and lesson_title not in seen:
                highlight_candidates.append(lesson_title)
                seen.add(lesson_title)
            if len(highlight_candidates) == 3:
                break
        label = f"W{week_index:02d}"
        weeks.append(
            CurriculumWeek(
                week_index=week_index,
                start_date=cursor,
                end_date=week_end,
                label=label,
                sessions=week_sessions,
                dominant_subject=dominant_subject,
                highlights=highlight_candidates,
            )
        )
        cursor = week_end + timedelta(days=1)
        week_index += 1

    curriculum = {
        "startDate": start.isoformat(),
        "endDate": end.isoformat(),
        "sessions": sessions,
        "weeks": [
            {
                "weekIndex": week.week_index,
                "label": week.label,
                "startDate": week.start_date.isoformat(),
                "endDate": week.end_date.isoformat(),
                "dominantSubject": week.dominant_subject,
                "highlights": week.highlights,
                "sessions": week.sessions,
            }
            for week in weeks
        ],
    }
    return curriculum, weeks


def week_for_date(target: date, weeks: list[CurriculumWeek]) -> int:
    for week in weeks:
        if week.start_date <= target <= week.end_date:
            return week.week_index
    if target < weeks[0].start_date:
        return 0
    return weeks[-1].week_index


def add_admission_data(students_by_name: dict[str, dict[str, Any]], course_start: date, weeks: list[CurriculumWeek]) -> None:
    wb = workbook_by_index(5)
    raw_ws = wb.worksheets[0]
    headers = [cell for cell in next(raw_ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    columns = find_header_indexes(
        headers,
        {
            "timestamp": ["타임스탬프", "wnwo"],
            "email": ["이메일"],
            "name": ["이름"],
            "intro": ["1분 자기소개"],
            "motivation": ["신청한 이유", "참여를 신청한 이유"],
            "experience": ["학습 및 업무 경험", "업무 경험"],
            "career": ["취업 분야"],
            "goal": ["수료 후 자신의 목표"],
            "conflict": ["갈등"],
            "failure": ["실수나 실패"],
            "peer": ["타인의 부족한 부분"],
        },
    )

    for row in raw_ws.iter_rows(min_row=2, values_only=True):
        name = clean_text(row[columns["name"]]) if columns["name"] >= 0 else ""
        student = students_by_name.get(normalize_name(name))
        if not student:
            continue
        submitted_at = parse_datetime(row[columns["timestamp"]]) if columns["timestamp"] >= 0 else None
        student["admission"] = {
            "submittedAt": submitted_at.isoformat() if submitted_at else "",
            "email": clean_text(row[columns["email"]]) if columns["email"] >= 0 else "",
            "intro": clean_text(row[columns["intro"]]) if columns["intro"] >= 0 else "",
            "motivation": clean_text(row[columns["motivation"]]) if columns["motivation"] >= 0 else "",
            "experience": clean_text(row[columns["experience"]]) if columns["experience"] >= 0 else "",
            "career": clean_text(row[columns["career"]]) if columns["career"] >= 0 else "",
            "goal": clean_text(row[columns["goal"]]) if columns["goal"] >= 0 else "",
            "conflict": clean_text(row[columns["conflict"]]) if columns["conflict"] >= 0 else "",
            "failure": clean_text(row[columns["failure"]]) if columns["failure"] >= 0 else "",
            "peer": clean_text(row[columns["peer"]]) if columns["peer"] >= 0 else "",
        }
        if submitted_at:
            student["timelineEvents"].append(
                {
                    "id": f"admission-{student['id']}",
                    "date": submitted_at.date().isoformat(),
                    "endDate": "",
                    "type": "admission",
                    "severity": "info",
                    "title": "지원서 제출",
                    "summary": short_text(student["admission"]["motivation"] or student["admission"]["intro"]),
                    "detail": student["admission"]["intro"] or student["admission"]["motivation"],
                    "projectPhase": "",
                    "sourceLabel": "모집 결과.xlsx",
                    "relatedWeek": week_for_date(submitted_at.date(), weeks),
                    "isEstimated": False,
                }
            )

    result_ws = wb.worksheets[2]
    for row in result_ws.iter_rows(min_row=11, values_only=True):
        name = clean_text(row[2])
        if not name:
            continue
        student = students_by_name.get(normalize_name(name))
        if not student:
            continue
        estimated_result_date = course_start - timedelta(days=1)
        result = clean_text(row[12])
        score = row[11] if row[11] is not None else ""
        total_comment = clean_text(row[17])
        student["admission"].update(
            {
                "interviewScore": score,
                "interviewResult": result,
                "interviewSummary": total_comment,
            }
        )
        student["timelineEvents"].append(
            {
                "id": f"admission-result-{student['id']}",
                "date": estimated_result_date.isoformat(),
                "endDate": "",
                "type": "admission",
                "severity": "success" if "합격" in result else "caution",
                "title": "최종 선발 결과",
                "summary": f"{result or '결과 확인 필요'} · 면접 점수 {score}",
                "detail": total_comment,
                "projectPhase": "",
                "sourceLabel": "모집 결과.xlsx",
                "relatedWeek": 0,
                "isEstimated": True,
            }
        )


def collect_project_date_sheet(workbook) -> list[date]:
    for sheet in workbook.worksheets:
        if "Date" == sheet.title:
            values = []
            for row in sheet.iter_rows(values_only=True):
                parsed = parse_date(row[0] if row else None)
                if parsed:
                    values.append(parsed)
            return sorted(set(values))
    return []


def build_project_phase_anchor_dates(
    phase_dates: dict[str, list[date]],
    extra_phases: list[str],
) -> dict[str, date]:
    anchors = {
        phase: min(dates)
        for phase, dates in phase_dates.items()
        if dates
    }
    ordered_phases = sorted(
        {phase for phase in list(anchors.keys()) + list(extra_phases)},
        key=phase_order_key,
    )
    numbered_anchors = sorted(
        (
            phase_order_key(phase),
            anchor_date,
        )
        for phase, anchor_date in anchors.items()
        if phase_order_key(phase) < 999
    )
    deltas = [
        (numbered_anchors[idx + 1][1] - numbered_anchors[idx][1]).days
        for idx in range(len(numbered_anchors) - 1)
        if 7 <= (numbered_anchors[idx + 1][1] - numbered_anchors[idx][1]).days <= 60
    ]
    default_gap = deltas[len(deltas) // 2] if deltas else 28

    for index, phase in enumerate(ordered_phases):
        if phase in anchors:
            continue
        prev_phase = next(
            (
                ordered_phases[cursor]
                for cursor in range(index - 1, -1, -1)
                if ordered_phases[cursor] in anchors
            ),
            None,
        )
        next_phase = next(
            (
                ordered_phases[cursor]
                for cursor in range(index + 1, len(ordered_phases))
                if ordered_phases[cursor] in anchors
            ),
            None,
        )
        if prev_phase and next_phase:
            prev_order = phase_order_key(prev_phase)
            next_order = phase_order_key(next_phase)
            distance = max(1, next_order - prev_order)
            step = max(7, (anchors[next_phase] - anchors[prev_phase]).days // distance)
            anchors[phase] = anchors[prev_phase] + timedelta(days=step * (phase_order_key(phase) - prev_order))
        elif prev_phase:
            anchors[phase] = anchors[prev_phase] + timedelta(days=default_gap * max(1, phase_order_key(phase) - phase_order_key(prev_phase)))
        elif next_phase:
            anchors[phase] = anchors[next_phase] - timedelta(days=default_gap * max(1, phase_order_key(next_phase) - phase_order_key(phase)))
    return anchors


def parse_project_team_history() -> dict[str, list[dict[str, Any]]]:
    path = DATA_DIR / "프로젝트 팀 구성.md"
    if not path.exists():
        return {}

    team_history_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    phase = ""
    team_label = ""
    team_members: list[dict[str, str]] = []

    def flush_team() -> None:
        nonlocal team_members
        if not phase or not team_label or not team_members:
            team_members = []
            return
        team_number_match = re.search(r"(\d+)", team_label)
        team_number = int(team_number_match.group(1)) if team_number_match else None
        for member in team_members:
            teammates = [
                {"name": peer["name"], "roleLabel": peer["roleLabel"]}
                for peer in team_members
                if peer["name"] != member["name"]
            ]
            role_label = member["roleLabel"]
            role = "team_lead" if "팀장" in role_label else "pm" if "pm" in role_label.lower() else "member"
            team_history_by_name[normalize_name(member["name"])].append(
                {
                    "phase": phase,
                    "teamLabel": team_label,
                    "teamNumber": team_number,
                    "role": role,
                    "roleLabel": role_label,
                    "teammates": teammates,
                }
            )
        team_members = []

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        phase_match = re.match(r"^##\s+(.+)$", line)
        if phase_match:
            flush_team()
            phase = phase_match.group(1).strip()
            team_label = ""
            continue
        team_match = re.match(r"^###\s+(.+)$", line)
        if team_match:
            flush_team()
            team_label = team_match.group(1).strip()
            continue
        member_match = re.match(r"^-?\s*([^\(\n]+?)\s*(?:\(([^)]+)\))?$", line)
        if member_match and team_label:
            name = clean_text(member_match.group(1))
            role_label = clean_text(member_match.group(2) or "팀원")
            if name:
                team_members.append({"name": name, "roleLabel": role_label})

    flush_team()
    for assignments in team_history_by_name.values():
        assignments.sort(key=lambda item: (phase_order_key(item["phase"]), item["teamNumber"] or 999))
    return team_history_by_name


def build_peer_relationships(team_history: list[dict[str, Any]]) -> list[dict[str, Any]]:
    relationship_map: dict[str, dict[str, Any]] = {}
    for assignment in team_history:
        for peer in assignment.get("teammates", []):
            key = normalize_name(peer["name"])
            relation = relationship_map.setdefault(
                key,
                {
                    "name": peer["name"],
                    "count": 0,
                    "phases": [],
                    "teams": [],
                },
            )
            relation["count"] += 1
            if assignment["phase"] not in relation["phases"]:
                relation["phases"].append(assignment["phase"])
            if assignment["teamLabel"] and assignment["teamLabel"] not in relation["teams"]:
                relation["teams"].append(assignment["teamLabel"])
    return sorted(
        relationship_map.values(),
        key=lambda item: (-item["count"], item["name"]),
    )


def parse_round_date(label: str) -> date | None:
    match = re.search(r"\((\d{2})(\d{2})(\d{2})\)", label)
    if not match:
        return None
    year = 2000 + int(match.group(1))
    month = int(match.group(2))
    day = int(match.group(3))
    return date(year, month, day)


def parse_career_documents() -> dict[str, dict[str, Any]]:
    base_dir = DATA_DIR / "이력서 및 자기소개"
    if not base_dir.exists():
        return {}

    records: dict[str, dict[str, Any]] = defaultdict(lambda: {"rounds": []})

    def ensure_round(student_name: str, round_label: str, round_date: date | None) -> dict[str, Any]:
        student_key = normalize_name(student_name)
        student_bucket = records[student_key]
        for entry in student_bucket["rounds"]:
            if entry["roundLabel"] == round_label:
                return entry
        entry = {
            "roundLabel": round_label,
            "roundIndex": phase_order_key(round_label),
            "date": round_date.isoformat() if round_date else "",
            "documents": {
                "selfIntroduction": "",
                "resume": "",
            },
            "feedback": "",
            "sourceFiles": [],
        }
        student_bucket["rounds"].append(entry)
        return entry

    for round_dir in sorted(path for path in base_dir.iterdir() if path.is_dir()):
        round_label = clean_text(round_dir.name).split("(")[0]
        round_date = parse_round_date(round_dir.name)

        if round_label == "1차":
            intro_dir = round_dir / "자소서 모음집"
            if intro_dir.exists():
                for path in sorted(intro_dir.glob("*.txt")):
                    if path.name.startswith("_"):
                        continue
                    student_name = clean_text(path.stem.split("_")[0])
                    round_entry = ensure_round(student_name, round_label, round_date)
                    round_entry["documents"]["selfIntroduction"] = clean_text(path.read_text(encoding="utf-8"))
                    round_entry["sourceFiles"].append(str(path.relative_to(DATA_DIR)).replace("\\", "/"))

            feedback_dir = round_dir / "개인 피드백"
            if feedback_dir.exists():
                for path in sorted(feedback_dir.glob("*.md")):
                    student_name = clean_text(path.stem)
                    round_entry = ensure_round(student_name, round_label, round_date)
                    round_entry["feedback"] = clean_text(path.read_text(encoding="utf-8"))
                    round_entry["sourceFiles"].append(str(path.relative_to(DATA_DIR)).replace("\\", "/"))

        if round_label == "2차":
            text_dir = round_dir / "피드백" / "_추출텍스트"
            if text_dir.exists():
                for path in sorted(text_dir.glob("*.txt")):
                    student_name = clean_text(path.stem.split("_")[0])
                    round_entry = ensure_round(student_name, round_label, round_date)
                    content = clean_text(path.read_text(encoding="utf-8"))
                    if "이력서" in path.stem:
                        round_entry["documents"]["resume"] = content
                    if "자기소개서" in path.stem:
                        round_entry["documents"]["selfIntroduction"] = content
                    round_entry["sourceFiles"].append(str(path.relative_to(DATA_DIR)).replace("\\", "/"))

            feedback_dir = round_dir / "피드백"
            if feedback_dir.exists():
                for path in sorted(feedback_dir.glob("*.md")):
                    if path.name.startswith("_"):
                        continue
                    student_name = clean_text(path.stem)
                    round_entry = ensure_round(student_name, round_label, round_date)
                    round_entry["feedback"] = clean_text(path.read_text(encoding="utf-8"))
                    round_entry["sourceFiles"].append(str(path.relative_to(DATA_DIR)).replace("\\", "/"))

    finalized: dict[str, dict[str, Any]] = {}
    for student_key, payload in records.items():
        rounds = sorted(payload["rounds"], key=lambda item: item["roundIndex"])
        self_intro_rounds = [item["roundLabel"] for item in rounds if item["documents"]["selfIntroduction"]]
        resume_rounds = [item["roundLabel"] for item in rounds if item["documents"]["resume"]]
        feedback_rounds = [item["roundLabel"] for item in rounds if item["feedback"]]
        latest_date = max((item["date"] for item in rounds if item["date"]), default="")
        finalized[student_key] = {
            "rounds": rounds,
            "summary": {
                "roundCount": len(rounds),
                "selfIntroductionRounds": self_intro_rounds,
                "resumeRounds": resume_rounds,
                "feedbackRounds": feedback_rounds,
                "latestDocumentDate": latest_date,
                "hasRevisionHistory": len(self_intro_rounds) >= 2,
            },
        }
    return finalized


def add_project_data(students_by_name: dict[str, dict[str, Any]], weeks: list[CurriculumWeek]) -> dict[str, list[date]]:
    course_start = weeks[0].start_date
    course_end = weeks[-1].end_date
    phase_dates: dict[str, list[date]] = {}
    for phase_meta in PROJECT_PHASES:
        phase = phase_meta["phase"]
        daily_wb = workbook_by_index(phase_meta["daily_index"])
        raw_ws = next(sheet for sheet in daily_wb.worksheets if "설문지 응답" in sheet.title)
        headers = list(next(raw_ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        columns = find_header_indexes(
            headers,
            {
                "timestamp": ["타임스탬프"],
                "name": ["이름"],
                "team": ["소속 팀", "소속팀"],
                "date": ["작성 날짜"],
                "workload": ["작업량", "작업 이력", "오늘 작업한 내용"],
                "note": ["자유로운 한마디"],
            },
        )

        known_dates = [
            item
            for item in collect_project_date_sheet(daily_wb)
            if date_in_window(item, course_start - timedelta(days=45), course_end + timedelta(days=30))
        ]

        for row in raw_ws.iter_rows(min_row=2, values_only=True):
            name = clean_text(row[columns["name"]]) if columns["name"] >= 0 else ""
            if not name:
                continue
            student = students_by_name.get(normalize_name(name))
            if not student:
                continue
            submitted_at = parse_datetime(row[columns["timestamp"]]) if columns["timestamp"] >= 0 else None
            assigned_date = normalize_project_date(
                row[columns["date"]] if columns["date"] >= 0 else None,
                submitted_at,
                course_start,
                course_end,
            )
            if assigned_date and assigned_date not in known_dates:
                known_dates.append(assigned_date)
            work_text = clean_text(row[columns["workload"]]) if columns["workload"] >= 0 else ""
            note_text = clean_text(row[columns["note"]]) if columns["note"] >= 0 else ""
            team = clean_text(row[columns["team"]]) if columns["team"] >= 0 else ""
            due_end = datetime.combine(assigned_date + timedelta(days=1), time(9, 0)) if assigned_date else None
            on_time = bool(submitted_at and due_end and submitted_at <= due_end)
            submission = {
                "phase": phase,
                "date": assigned_date.isoformat() if assigned_date else "",
                "submittedAt": submitted_at.isoformat() if submitted_at else "",
                "team": team,
                "workText": work_text,
                "noteText": note_text,
                "onTime": on_time,
            }
            student["checkins"].append(submission)
            if assigned_date:
                student["timelineEvents"].append(
                    {
                        "id": f"{phase}-{student['id']}-{assigned_date.isoformat()}",
                        "date": assigned_date.isoformat(),
                        "endDate": "",
                        "type": "project",
                        "severity": "info" if on_time else "caution",
                        "title": f"{phase} 데일리 체크인",
                        "summary": short_text(work_text or note_text or "프로젝트 일일 기록 제출"),
                        "detail": "\n\n".join(filter(None, [work_text, note_text])),
                        "projectPhase": phase,
                        "sourceLabel": "프로젝트 데일리 체크인",
                        "relatedWeek": week_for_date(assigned_date, weeks),
                        "isEstimated": False,
                    }
                )

        known_dates = dominant_phase_dates(known_dates)
        phase_dates[phase] = known_dates

        retro_index = phase_meta["retro_index"]
        if retro_index is not None:
            retro_wb = workbook_by_index(retro_index)
            retro_ws = retro_wb.worksheets[0]
            retro_headers = list(next(retro_ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            retro_cols = find_header_indexes(
                retro_headers,
                {
                    "timestamp": ["타임스탬프"],
                    "name": ["이름"],
                    "team": ["소속팀", "소속 팀"],
                },
            )
            text_indexes = [
                idx
                for idx, header in enumerate(retro_headers)
                if idx > 3 and compact_text(header)
            ]
            for row in retro_ws.iter_rows(min_row=2, values_only=True):
                name = clean_text(row[retro_cols["name"]]) if retro_cols["name"] >= 0 else ""
                if not name:
                    continue
                student = students_by_name.get(normalize_name(name))
                if not student:
                    continue
                submitted_at = parse_datetime(row[retro_cols["timestamp"]]) if retro_cols["timestamp"] >= 0 else None
                retro_date = submitted_at.date() if submitted_at else (known_dates[-1] if known_dates else None)
                text_blocks = []
                for idx in text_indexes:
                    header = compact_text(retro_headers[idx])
                    content = clean_text(row[idx])
                    if content:
                        text_blocks.append(f"{header}\n{content}")
                retro_text = "\n\n".join(text_blocks)
                student["retrospectives"].append(
                    {
                        "phase": phase,
                        "date": retro_date.isoformat() if retro_date else "",
                        "submittedAt": submitted_at.isoformat() if submitted_at else "",
                        "detail": retro_text,
                    }
                )
                if retro_date:
                    student["timelineEvents"].append(
                        {
                            "id": f"{phase}-retro-{student['id']}-{retro_date.isoformat()}",
                            "date": retro_date.isoformat(),
                            "endDate": "",
                            "type": "retro",
                            "severity": "info",
                            "title": f"{phase} 회고 제출",
                            "summary": short_text(retro_text or "프로젝트 종료 후 회고 제출"),
                            "detail": retro_text,
                            "projectPhase": phase,
                            "sourceLabel": "프로젝트 회고",
                            "relatedWeek": week_for_date(retro_date, weeks),
                            "isEstimated": False,
                        }
                    )

    return phase_dates


def add_project_team_data(
    students_by_name: dict[str, dict[str, Any]],
    weeks: list[CurriculumWeek],
    phase_dates: dict[str, list[date]],
) -> None:
    team_history_map = parse_project_team_history()
    phase_anchors = build_project_phase_anchor_dates(
        phase_dates,
        sorted(
            {
                assignment["phase"]
                for assignments in team_history_map.values()
                for assignment in assignments
            },
            key=phase_order_key,
        ),
    )

    for student_key, assignments in team_history_map.items():
        student = students_by_name.get(student_key)
        if not student:
            continue
        student_assignments = []
        for assignment in assignments:
            event_date = phase_anchors.get(assignment["phase"])
            student_assignments.append(
                {
                    **assignment,
                    "date": event_date.isoformat() if event_date else "",
                }
            )
        student["projectTeamHistory"] = student_assignments
        student["peerRelationships"] = build_peer_relationships(student_assignments)
        for index, assignment in enumerate(student_assignments, start=1):
            event_date = parse_date(assignment.get("date"))
            if not event_date:
                continue
            role_label = assignment.get("roleLabel") or "팀원"
            teammate_names = [peer["name"] for peer in assignment.get("teammates", [])]
            student["timelineEvents"].append(
                {
                    "id": f"team-{student['id']}-{phase_order_key(assignment['phase'])}-{index}",
                    "date": event_date.isoformat(),
                    "endDate": "",
                    "type": "project",
                    "severity": "info",
                    "title": f"{assignment['phase']} 팀 배치",
                    "summary": f"{assignment['teamLabel']} · 역할 {role_label} · 팀원 {len(teammate_names)}명",
                    "detail": "\n".join(
                        [
                            f"팀: {assignment['teamLabel']}",
                            f"역할: {role_label}",
                            f"함께한 팀원: {', '.join(teammate_names) if teammate_names else '-'}",
                        ]
                    ),
                    "projectPhase": assignment["phase"],
                    "sourceLabel": "프로젝트 팀 구성",
                    "relatedWeek": week_for_date(event_date, weeks),
                    "isEstimated": True,
                }
            )


def add_career_document_data(students_by_name: dict[str, dict[str, Any]], weeks: list[CurriculumWeek]) -> None:
    career_documents_map = parse_career_documents()
    for student_key, payload in career_documents_map.items():
        student = students_by_name.get(student_key)
        if not student:
            continue
        student["careerDocuments"] = payload
        for round_entry in payload.get("rounds", []):
            event_date = parse_date(round_entry.get("date"))
            if not event_date:
                continue
            round_label = round_entry["roundLabel"]
            documents = round_entry.get("documents", {})
            if documents.get("selfIntroduction"):
                student["timelineEvents"].append(
                    {
                        "id": f"career-intro-{student['id']}-{round_label}",
                        "date": event_date.isoformat(),
                        "endDate": "",
                        "type": "career",
                        "severity": "info",
                        "title": f"{round_label} 자기소개서 작성",
                        "summary": short_text(documents["selfIntroduction"]),
                        "detail": documents["selfIntroduction"],
                        "projectPhase": "취업 준비",
                        "sourceLabel": "이력서 및 자기소개",
                        "relatedWeek": week_for_date(event_date, weeks),
                        "isEstimated": False,
                    }
                )
            if documents.get("resume"):
                student["timelineEvents"].append(
                    {
                        "id": f"career-resume-{student['id']}-{round_label}",
                        "date": event_date.isoformat(),
                        "endDate": "",
                        "type": "career",
                        "severity": "info",
                        "title": f"{round_label} 이력서 작성",
                        "summary": short_text(documents["resume"]),
                        "detail": documents["resume"],
                        "projectPhase": "취업 준비",
                        "sourceLabel": "이력서 및 자기소개",
                        "relatedWeek": week_for_date(event_date, weeks),
                        "isEstimated": False,
                    }
                )
            if round_entry.get("feedback"):
                student["timelineEvents"].append(
                    {
                        "id": f"career-feedback-{student['id']}-{round_label}",
                        "date": event_date.isoformat(),
                        "endDate": "",
                        "type": "career",
                        "severity": "info",
                        "title": f"{round_label} 취업 문서 피드백",
                        "summary": short_text(round_entry["feedback"]),
                        "detail": round_entry["feedback"],
                        "projectPhase": "취업 준비",
                        "sourceLabel": "이력서 및 자기소개",
                        "relatedWeek": week_for_date(event_date, weeks),
                        "isEstimated": False,
                    }
                )


def add_attendance_data(students_by_name: dict[str, dict[str, Any]], weeks: list[CurriculumWeek]) -> None:
    wb = workbook_by_index(7)
    for ws in wb.worksheets:
        if ws.title in {"지각", "결석"}:
            continue
        headers = list(next(ws.iter_rows(min_row=4, max_row=4, values_only=True)))
        date_columns: list[tuple[int, date]] = []
        for idx, value in enumerate(headers):
            parsed = parse_date(value)
            if parsed:
                date_columns.append((idx, parsed))
        for row in ws.iter_rows(min_row=5, values_only=True):
            name = clean_text(row[0])
            phone = clean_text(row[1])
            if not name:
                continue
            student = students_by_name.get(normalize_name(name))
            if not student:
                continue
            if phone and student["phone"] and phone != student["phone"]:
                continue
            for idx, event_date in date_columns:
                raw_value = clean_text(row[idx]) if idx < len(row) else ""
                if not raw_value:
                    continue
                kind = attendance_kind(raw_value)
                reason_profile = attendance_reason_profile(raw_value, kind)
                severity = reason_profile["severity"]
                event = {
                    "id": f"attendance-{student['id']}-{event_date.isoformat()}-{idx}",
                    "date": event_date.isoformat(),
                    "endDate": "",
                    "type": "attendance",
                    "severity": severity,
                    "title": f"출결 이슈 · {kind}",
                    "summary": short_text(raw_value),
                    "detail": raw_value,
                    "projectPhase": "",
                    "sourceLabel": "출결 기입 시트",
                    "relatedWeek": week_for_date(event_date, weeks),
                    "isEstimated": False,
                    "attendanceCategory": reason_profile["category"],
                    "scoreImpact": reason_profile["impact"],
                }
                student["attendanceEvents"].append(
                    {
                        "date": event_date.isoformat(),
                        "kind": kind,
                        "detail": raw_value,
                        "severity": severity,
                        "category": reason_profile["category"],
                        "impact": reason_profile["impact"],
                    }
                )
                student["timelineEvents"].append(event)


def parse_counseling_sections(text: str) -> list[dict[str, str]]:
    sections: list[dict[str, str]] = []
    pattern = re.compile(r"^##\s+(.+)$", re.MULTILINE)
    matches = list(pattern.finditer(text))
    if not matches:
        return sections
    for idx, match in enumerate(matches):
        title = match.group(1).strip()
        start = match.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
        block = text[start:end].strip()
        counselor_match = re.search(r"-\s*면담자\s*:\s*(.+)", block)
        date_match = re.search(r"-\s*면담일\s*:\s*(\d{4}-\d{2}-\d{2})", block)
        content = re.sub(r"-\s*면담자\s*:\s*.+", "", block)
        content = re.sub(r"-\s*면담일\s*:\s*.+", "", content)
        content = content.strip()
        sections.append(
            {
                "title": title,
                "counselor": counselor_match.group(1).strip() if counselor_match else "",
                "date": date_match.group(1).strip() if date_match else "",
                "content": content,
            }
        )
    return sections


def add_counseling_data(students_by_name: dict[str, dict[str, Any]], weeks: list[CurriculumWeek]) -> None:
    counsel_dir = DATA_DIR / "면담 기록"
    for path in sorted(counsel_dir.glob("*.md")):
        student = students_by_name.get(normalize_name(path.stem))
        if not student:
            continue
        text = path.read_text(encoding="utf-8")
        sections = parse_counseling_sections(text)
        for index, section in enumerate(sections, start=1):
            event_date = parse_date(section["date"])
            severity = "warning" if ("경고" in section["title"] or "이탈" in section["title"]) else "info"
            summary = short_text(section["content"])
            counseling = {
                "title": section["title"],
                "counselor": section["counselor"],
                "date": event_date.isoformat() if event_date else "",
                "content": section["content"],
                "severity": severity,
            }
            student["counselings"].append(counseling)
            if event_date:
                student["timelineEvents"].append(
                    {
                        "id": f"counseling-{student['id']}-{event_date.isoformat()}-{index}",
                        "date": event_date.isoformat(),
                        "endDate": "",
                        "type": "counseling",
                        "severity": severity,
                        "title": section["title"],
                        "summary": summary,
                        "detail": section["content"],
                        "projectPhase": "",
                        "sourceLabel": "면담 기록",
                        "relatedWeek": week_for_date(event_date, weeks),
                        "isEstimated": False,
                    }
                )


def build_snapshot(student: dict[str, Any], label: str, snapshot_date: date, cutoff: date, phase_dates: dict[str, list[date]]) -> dict[str, Any]:
    attendance_events = [event for event in student["attendanceEvents"] if parse_date(event["date"]) and parse_date(event["date"]) <= cutoff]
    counselings = [item for item in student["counselings"] if parse_date(item["date"]) and parse_date(item["date"]) <= cutoff]
    checkins = [item for item in student["checkins"] if parse_date(item["date"]) and parse_date(item["date"]) <= cutoff]
    retros = [item for item in student["retrospectives"] if parse_date(item["date"]) and parse_date(item["date"]) <= cutoff]
    team_history = [item for item in student.get("projectTeamHistory", []) if parse_date(item.get("date")) is None or parse_date(item.get("date")) <= cutoff]
    career_rounds = [item for item in student.get("careerDocuments", {}).get("rounds", []) if not item.get("date") or parse_date(item.get("date")) <= cutoff]
    admission = student.get("admission", {})

    expected_dates = []
    for dates in phase_dates.values():
        expected_dates.extend([d for d in dates if d <= cutoff])
    expected_dates = sorted(set(expected_dates))

    submission_dates = {parse_date(item["date"]) for item in checkins if parse_date(item["date"])}
    response_rate = len(submission_dates) / len(expected_dates) if expected_dates else 0.0
    behavioral_attendance_issues = [
        event for event in attendance_events if event.get("impact") == "behavioral"
    ]
    contextual_attendance_issues = [
        event for event in attendance_events if event.get("impact") == "contextual"
    ]
    exempt_attendance_issues = [
        event for event in attendance_events if event.get("impact") == "exempt"
    ]
    health_attendance_issues = [
        event for event in attendance_events if event.get("category") == "health"
    ]
    risk_attendance_count = len(behavioral_attendance_issues)
    behavioral_absence_count = sum(
        1 for event in behavioral_attendance_issues if event["kind"] == "결석"
    )
    on_time_rate = (
        sum(1 for item in checkins if item["onTime"]) / len(checkins)
        if checkins
        else 0.0
    )

    all_text = "\n".join(
        [
            admission.get("intro", ""),
            admission.get("motivation", ""),
            admission.get("goal", ""),
            admission.get("conflict", ""),
            admission.get("failure", ""),
            admission.get("peer", ""),
            "\n".join(item["detail"] for item in retros),
            "\n".join(item["content"] for item in counselings),
            "\n".join(item["workText"] + "\n" + item["noteText"] for item in checkins),
            "\n".join(
                "\n".join(
                    filter(
                        None,
                        [
                            round_item.get("documents", {}).get("selfIntroduction", ""),
                            round_item.get("documents", {}).get("resume", ""),
                            round_item.get("feedback", ""),
                        ],
                    )
                )
                for round_item in career_rounds
            ),
        ]
    )

    positive_collab = keyword_hits(all_text, POSITIVE_COLLAB_KEYWORDS)
    negative_collab = keyword_hits(all_text, NEGATIVE_COLLAB_KEYWORDS)
    resilience_up = keyword_hits(all_text, RESILIENCE_KEYWORDS)
    resilience_down = keyword_hits(all_text, LOW_RESILIENCE_KEYWORDS)
    reflection_hits = keyword_hits(all_text, REFLECTION_KEYWORDS)
    career_hits = keyword_hits(all_text, CAREER_KEYWORDS)
    leader_count = sum(1 for item in team_history if item.get("role") in {"team_lead", "pm"})
    feedback_round_count = sum(1 for item in career_rounds if item.get("feedback"))
    resume_round_count = sum(1 for item in career_rounds if item.get("documents", {}).get("resume"))
    self_intro_round_count = sum(1 for item in career_rounds if item.get("documents", {}).get("selfIntroduction"))
    has_revision_history = self_intro_round_count >= 2

    self_regulation = 2.0
    if response_rate >= 0.8:
        self_regulation += 1
    if response_rate >= 0.95 and on_time_rate >= 0.85:
        self_regulation += 0.5
    if response_rate < 0.6:
        self_regulation -= 0.8
    if risk_attendance_count >= 3:
        self_regulation -= 0.8
    if behavioral_absence_count >= 1:
        self_regulation -= 0.5
    if has_revision_history:
        self_regulation += 0.2

    engagement = 2.0
    if response_rate >= 0.75 and risk_attendance_count <= 1:
        engagement += 1
    if response_rate >= 0.9 and behavioral_absence_count == 0:
        engagement += 0.5
    if response_rate < 0.5:
        engagement -= 0.9
    if risk_attendance_count >= 2:
        engagement -= 0.5
    if behavioral_absence_count >= 1:
        engagement -= 0.7

    collaboration = 2.0
    if positive_collab > negative_collab:
        collaboration += 0.8
    if positive_collab >= 5:
        collaboration += 0.5
    if negative_collab >= 4:
        collaboration -= 0.8
    if any("경고" in counseling["title"] or "이탈" in counseling["title"] for counseling in counselings):
        collaboration -= 0.6
    if leader_count >= 1:
        collaboration += 0.3
    if leader_count >= 2:
        collaboration += 0.2

    resilience = 2.0
    if resilience_up > resilience_down:
        resilience += 0.8
    if reflection_hits >= 5:
        resilience += 0.3
    if resilience_down >= 4:
        resilience -= 0.9
    if behavioral_absence_count >= 1:
        resilience -= 0.4

    reflection = 2.0
    if reflection_hits >= 4:
        reflection += 1
    if len(retros) >= 2:
        reflection += 0.5
    if not retros and not counselings:
        reflection -= 0.5
    if feedback_round_count >= 1:
        reflection += 0.2
    if has_revision_history:
        reflection += 0.2

    career_agency = 2.0
    if len(admission.get("goal", "")) > 60 or len(admission.get("career", "")) > 40:
        career_agency += 0.8
    if career_hits >= 5:
        career_agency += 0.7
    if "명확한 목표는 없어" in all_text or "모르겠" in all_text:
        career_agency -= 0.7
    if self_intro_round_count >= 1:
        career_agency += 0.5
    if resume_round_count >= 1:
        career_agency += 0.4
    if feedback_round_count >= 1:
        career_agency += 0.2

    snapshot_scores = {
        "selfRegulation": clamp_score(self_regulation),
        "engagement": clamp_score(engagement),
        "collaboration": clamp_score(collaboration),
        "resilience": clamp_score(resilience),
        "reflection": clamp_score(reflection),
        "careerAgency": clamp_score(career_agency),
    }

    sources = 0
    for value in [attendance_events, counselings, checkins, retros, team_history, career_rounds]:
        if value:
            sources += 1
    if admission:
        sources += 1
    if len(exempt_attendance_issues) >= 2 and risk_attendance_count == 0:
        confidence = "High" if sources >= 4 else "Medium" if sources >= 2 else "Low"
    else:
        confidence = "High" if sources >= 4 else "Medium" if sources >= 2 else "Low"

    weakest_area = min(snapshot_scores.items(), key=lambda item: item[1])[0]
    weakest_map = {
        "selfRegulation": "자기조절 지원 필요",
        "engagement": "참여 지속성 점검 필요",
        "collaboration": "협업 상황 면담 권장",
        "resilience": "도전 대응 지원 필요",
        "reflection": "회고 구조화 필요",
        "careerAgency": "진로 목적성 구체화 필요",
    }

    note = weakest_map[weakest_area]
    if max(snapshot_scores.values()) >= 4 and min(snapshot_scores.values()) >= 3:
        note = "전반적으로 안정적인 성장 흐름"
    elif len(health_attendance_issues) >= 2 and risk_attendance_count == 0:
        note = "건강 및 컨디션 변동 맥락을 함께 살필 필요"
    elif len(exempt_attendance_issues) >= 2 and risk_attendance_count == 0:
        note = "행정 및 외부 일정이 반복되어 맥락 확인 필요"
    elif len(contextual_attendance_issues) >= 2 and risk_attendance_count == 0 and weakest_area == "engagement":
        note = "출결 기록은 있으나 성향 감점보다 배경 확인이 우선"

    return {
        "snapshotId": f"{student['id']}-{label}",
        "snapshotDate": snapshot_date.isoformat(),
        "snapshotType": label,
        "scores": snapshot_scores,
        "confidence": confidence,
        "note": note,
    }


def build_evaluation_and_status(students: list[dict[str, Any]], curriculum: dict[str, Any], weeks: list[CurriculumWeek], phase_dates: dict[str, list[date]]) -> None:
    course_start = parse_date(curriculum["startDate"])
    current_end = parse_date(curriculum["endDate"])
    if not course_start or not current_end:
        return

    snapshot_targets = [
        ("입과 초기", course_start),
        ("1차 프로젝트 이후", max(phase_dates.get("1차 프로젝트", [course_start])) + timedelta(days=3) if phase_dates.get("1차 프로젝트") else course_start),
        ("2차 프로젝트 이후", max(phase_dates.get("2차 프로젝트", [course_start])) + timedelta(days=3) if phase_dates.get("2차 프로젝트") else course_start + timedelta(days=60)),
        ("현재", current_end),
    ]

    for student in students:
        snapshots = []
        for label, cutoff in snapshot_targets:
            snapshots.append(build_snapshot(student, label, cutoff, cutoff, phase_dates))
        student["evaluationSnapshots"] = snapshots
        latest_scores = snapshots[-1]["scores"]
        student["currentProfile"] = {
            **latest_scores,
            "confidence": snapshots[-1]["confidence"],
            "note": snapshots[-1]["note"],
            "updatedAt": snapshots[-1]["snapshotDate"],
        }

        weekly_map: dict[int, dict[str, Any]] = {}
        for week in weeks:
            weekly_map[week.week_index] = {
                "weekIndex": week.week_index,
                "label": week.label,
                "startDate": week.start_date.isoformat(),
                "endDate": week.end_date.isoformat(),
                "counts": {
                    "attendance": 0,
                    "attendanceRisk": 0,
                    "project": 0,
                    "retro": 0,
                    "counseling": 0,
                    "admission": 0,
                    "career": 0,
                },
                "severity": "stable",
                "headline": "안정적인 학습 흐름",
                "notes": [],
                "snapshot": None,
            }
        weekly_map[0] = {
            "weekIndex": 0,
            "label": "모집",
            "startDate": "",
            "endDate": curriculum["startDate"],
            "counts": {
                "attendance": 0,
                "attendanceRisk": 0,
                "project": 0,
                "retro": 0,
                "counseling": 0,
                "admission": 0,
                "career": 0,
            },
            "severity": "stable",
            "headline": "모집 및 선발 기록",
            "notes": [],
            "snapshot": None,
        }

        for event in sorted(student["timelineEvents"], key=lambda item: (item["date"], item["title"])):
            week_index = event.get("relatedWeek", 0)
            if week_index not in weekly_map:
                continue
            week = weekly_map[week_index]
            event_type = event["type"]
            if event_type == "attendance":
                week["counts"]["attendance"] += 1
                if event.get("scoreImpact") == "behavioral":
                    week["counts"]["attendanceRisk"] += 1
            elif event_type in week["counts"]:
                week["counts"][event_type] += 1
            elif event_type == "project":
                week["counts"]["project"] += 1
            elif event_type == "retro":
                week["counts"]["retro"] += 1
            elif event_type == "counseling":
                week["counts"]["counseling"] += 1
            elif event_type == "admission":
                week["counts"]["admission"] += 1
            elif event_type == "career":
                week["counts"]["career"] += 1
            week["notes"].append(
                {
                    "type": event["type"],
                    "title": event["title"],
                    "summary": event["summary"],
                    "severity": event["severity"],
                    "date": event["date"],
                }
            )
            if severity_rank(event["severity"]) >= severity_rank(week["severity"]):
                week["severity"] = event["severity"] if event["severity"] != "success" else "stable"

        for snapshot in snapshots:
            snapshot_week = week_for_date(parse_date(snapshot["snapshotDate"]), weeks)
            if snapshot_week in weekly_map:
                weekly_map[snapshot_week]["snapshot"] = snapshot
                weekly_map[snapshot_week]["notes"].append(
                    {
                        "type": "snapshot",
                        "title": "성향 스냅샷 갱신",
                        "summary": snapshot["note"],
                        "severity": "info",
                        "date": snapshot["snapshotDate"],
                    }
                )

        # derive headlines
        for week_index, week in weekly_map.items():
            attendance_count = week["counts"]["attendance"]
            attendance_risk_count = week["counts"]["attendanceRisk"]
            counseling_count = week["counts"]["counseling"]
            project_count = week["counts"]["project"]
            retro_count = week["counts"]["retro"]
            career_count = week["counts"]["career"]
            if week["severity"] == "warning":
                week["headline"] = "집중 확인이 필요한 구간"
            elif attendance_risk_count >= 2:
                week["headline"] = "출결 위험 신호가 집중된 구간"
                week["severity"] = "caution"
            elif attendance_risk_count == 1:
                week["headline"] = "출결 위험 신호를 확인한 구간"
                week["severity"] = "caution"
            elif attendance_count >= 2:
                week["headline"] = "출결 배경 확인이 필요한 구간"
            elif counseling_count >= 1:
                week["headline"] = "면담 개입이 있었던 구간"
            elif career_count >= 1:
                week["headline"] = "취업 문서 학습이 진행된 구간"
            elif project_count >= 2 or retro_count >= 1:
                week["headline"] = "프로젝트 활동이 밀집된 구간"
            elif week_index == 0:
                week["headline"] = "모집 및 선발 단계"

            week["notes"] = sorted(
                week["notes"],
                key=lambda item: (severity_rank(item["severity"]), item["date"]),
                reverse=True,
            )[:4]

        # status periods from severity streaks
        active_period = None
        periods = []
        for week_index in sorted(index for index in weekly_map.keys() if index != 0):
            week = weekly_map[week_index]
            if week["severity"] in {"caution", "warning"}:
                if active_period is None:
                    active_period = {
                        "statusType": "집중 관찰" if week["severity"] == "warning" else "주의",
                        "severity": week["severity"],
                        "startDate": week["startDate"],
                        "endDate": week["endDate"],
                        "reasonSummary": week["headline"],
                    }
                else:
                    active_period["endDate"] = week["endDate"]
                    if week["severity"] == "warning":
                        active_period["statusType"] = "경고"
                        active_period["severity"] = "warning"
            elif active_period is not None:
                periods.append(active_period)
                active_period = None
        if active_period is not None:
            periods.append(active_period)

        student["statusPeriods"] = [
            {
                "statusPeriodId": f"{student['id']}-status-{idx + 1}",
                **period,
            }
            for idx, period in enumerate(periods)
        ]
        student["weeklyTimeline"] = [weekly_map[0]] + [weekly_map[week.week_index] for week in weeks]

        attendance_total = len(student["attendanceEvents"])
        attendance_risk_total = len(
            [event for event in student["attendanceEvents"] if event.get("impact") == "behavioral"]
        )
        project_expected_total = sum(len(dates) for dates in phase_dates.values())
        project_submissions = len({(item["phase"], item["date"]) for item in student["checkins"] if item["date"]})
        counseling_dates = [item["date"] for item in student["counselings"] if item["date"]]
        student["stats"] = {
            "attendanceIssues": attendance_total,
            "attendanceRiskIssues": attendance_risk_total,
            "lateCount": sum(1 for event in student["attendanceEvents"] if event["kind"] == "지각"),
            "absenceCount": sum(1 for event in student["attendanceEvents"] if event["kind"] == "결석"),
            "counselingCount": len(student["counselings"]),
            "projectSubmissionRate": round((project_submissions / project_expected_total) * 100, 1) if project_expected_total else 0,
            "latestCounselingDate": max(counseling_dates) if counseling_dates else "",
            "leadershipRoleCount": sum(1 for item in student.get("projectTeamHistory", []) if item.get("role") in {"team_lead", "pm"}),
            "repeatedPeerCount": len([item for item in student.get("peerRelationships", []) if item.get("count", 0) >= 2]),
            "careerDocumentRounds": student.get("careerDocuments", {}).get("summary", {}).get("roundCount", 0),
            "currentStatus": student["statusPeriods"][-1]["statusType"] if student["statusPeriods"] else "안정",
        }

        student["timelineEvents"] = sorted(
            student["timelineEvents"],
            key=lambda item: (item["date"], severity_rank(item["severity"])),
        )


def event_in_range(event_date: str, start: date, end: date) -> bool:
    parsed = parse_date(event_date)
    return bool(parsed and start <= parsed <= end)


def build_project_phase_ranges(curriculum: dict[str, Any], phase_dates: dict[str, list[date]]) -> list[dict[str, Any]]:
    course_start = parse_date(curriculum["startDate"])
    course_end = parse_date(curriculum["endDate"])
    if not course_start or not course_end:
        return []

    phase_end_dates = {
        label: max(
            [
                item
                for item in phase_dates.get(label, [])
                if date_in_window(item, course_start - timedelta(days=45), course_end + timedelta(days=30))
            ],
            default=None,
        )
        for label in EXPECTED_PROJECT_PHASE_LABELS
    }

    for index, label in enumerate(EXPECTED_PROJECT_PHASE_LABELS):
        if phase_end_dates[label]:
            continue

        previous_labels = EXPECTED_PROJECT_PHASE_LABELS[:index]
        next_labels = EXPECTED_PROJECT_PHASE_LABELS[index + 1 :]
        previous_end = next(
            (phase_end_dates[item] for item in reversed(previous_labels) if phase_end_dates[item]),
            None,
        )
        next_end = next(
            (phase_end_dates[item] for item in next_labels if phase_end_dates[item]),
            None,
        )

        if previous_end and next_end:
            prev_index = max(i for i, item in enumerate(EXPECTED_PROJECT_PHASE_LABELS[:index]) if phase_end_dates[item])
            next_index = min(
                index + 1 + i
                for i, item in enumerate(EXPECTED_PROJECT_PHASE_LABELS[index + 1 :])
                if phase_end_dates[item]
            )
            step = max(7, (next_end - previous_end).days // max(1, next_index - prev_index))
            phase_end_dates[label] = previous_end + timedelta(days=step * (index - prev_index))
        elif previous_end:
            remaining_segments = (len(EXPECTED_PROJECT_PHASE_LABELS) - index) + 1
            remaining_days = max(14, (course_end - previous_end).days)
            step = max(7, remaining_days // remaining_segments)
            phase_end_dates[label] = min(course_end - timedelta(days=remaining_segments - 1), previous_end + timedelta(days=step))
        elif next_end:
            remaining_segments = index + 1
            remaining_days = max(14, (next_end - course_start).days)
            step = max(7, remaining_days // remaining_segments)
            phase_end_dates[label] = course_start + timedelta(days=step * (index + 1))
        else:
            total_segments = len(EXPECTED_PROJECT_PHASE_LABELS) + 1
            total_days = max(28, (course_end - course_start).days)
            step = max(7, total_days // total_segments)
            phase_end_dates[label] = course_start + timedelta(days=step * (index + 1))

    ranges = []
    range_start = course_start
    for label in EXPECTED_PROJECT_PHASE_LABELS:
        range_end = min(course_end, phase_end_dates[label] or course_end)
        ranges.append(
            {
                "phase": label,
                "startDate": range_start.isoformat(),
                "endDate": range_end.isoformat(),
                "dates": [item.isoformat() for item in sorted(phase_dates.get(label, [])) if item],
                "isEstimated": not bool(phase_dates.get(label)),
            }
        )
        range_start = min(course_end, range_end + timedelta(days=1))
    return ranges


def build_milestones(
    curriculum: dict[str, Any],
    phase_ranges: list[dict[str, Any]],
    students: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    course_start = parse_date(curriculum["startDate"])
    course_end = parse_date(curriculum["endDate"])
    if not course_start or not course_end:
        return []

    admission_dates = [
        parse_date(event["date"])
        for student in students
        for event in student.get("timelineEvents", [])
        if event.get("type") == "admission" and parse_date(event.get("date"))
    ]
    admission_start = min(admission_dates, default=course_start - timedelta(days=21))
    phase_ends = {item["phase"]: parse_date(item["endDate"]) for item in phase_ranges}

    milestone_specs = [
        {
            "id": "m1",
            "label": "모집 ~ 개강",
            "startDate": admission_start.isoformat(),
            "endDate": (course_start - timedelta(days=1)).isoformat(),
            "isEstimated": False,
        },
        {
            "id": "m2",
            "label": "개강 ~ 1차 프로젝트 종료",
            "startDate": course_start.isoformat(),
            "endDate": (phase_ends.get(EXPECTED_PROJECT_PHASE_LABELS[0]) or course_start).isoformat(),
            "isEstimated": phase_ranges[0]["isEstimated"] if phase_ranges else False,
        },
        {
            "id": "m3",
            "label": "1차 프로젝트 종료 ~ 2차 프로젝트 종료",
            "startDate": ((phase_ends.get(EXPECTED_PROJECT_PHASE_LABELS[0]) or course_start) + timedelta(days=1)).isoformat(),
            "endDate": (phase_ends.get(EXPECTED_PROJECT_PHASE_LABELS[1]) or course_start).isoformat(),
            "isEstimated": phase_ranges[1]["isEstimated"] if len(phase_ranges) > 1 else False,
        },
        {
            "id": "m4",
            "label": "2차 프로젝트 종료 ~ 3차 프로젝트 종료",
            "startDate": ((phase_ends.get(EXPECTED_PROJECT_PHASE_LABELS[1]) or course_start) + timedelta(days=1)).isoformat(),
            "endDate": (phase_ends.get(EXPECTED_PROJECT_PHASE_LABELS[2]) or course_start).isoformat(),
            "isEstimated": phase_ranges[2]["isEstimated"] if len(phase_ranges) > 2 else False,
        },
        {
            "id": "m5",
            "label": "3차 프로젝트 종료 ~ 4차 프로젝트 종료",
            "startDate": ((phase_ends.get(EXPECTED_PROJECT_PHASE_LABELS[2]) or course_start) + timedelta(days=1)).isoformat(),
            "endDate": (phase_ends.get(EXPECTED_PROJECT_PHASE_LABELS[3]) or course_end).isoformat(),
            "isEstimated": phase_ranges[3]["isEstimated"] if len(phase_ranges) > 3 else True,
        },
        {
            "id": "m6",
            "label": "4차 프로젝트 종료 ~ 종강",
            "startDate": ((phase_ends.get(EXPECTED_PROJECT_PHASE_LABELS[3]) or course_end) + timedelta(days=1)).isoformat(),
            "endDate": course_end.isoformat(),
            "isEstimated": phase_ranges[3]["isEstimated"] if len(phase_ranges) > 3 else True,
        },
    ]

    milestones = []
    for item in milestone_specs:
        start_date = parse_date(item["startDate"]) or course_start
        end_date = parse_date(item["endDate"]) or course_end
        if end_date < start_date:
            end_date = start_date
        milestones.append(
            {
                **item,
                "startDate": start_date.isoformat(),
                "endDate": end_date.isoformat(),
            }
        )
    return milestones


def summarize_milestone(
    student: dict[str, Any],
    milestone: dict[str, Any],
    phase_dates: dict[str, list[date]],
    previous_average: float | None,
) -> dict[str, Any]:
    start_date = parse_date(milestone["startDate"])
    end_date = parse_date(milestone["endDate"])
    if not start_date or not end_date:
        return {}

    snapshot = build_snapshot(student, milestone["label"], end_date, end_date, phase_dates)
    profile_avg = profile_average(snapshot["scores"])
    events = [
        event
        for event in student.get("timelineEvents", [])
        if event_in_range(event.get("date", ""), start_date, end_date)
    ]
    role_history = [
        item
        for item in student.get("projectTeamHistory", [])
        if event_in_range(item.get("date", ""), start_date, end_date)
    ]
    career_rounds = [
        item
        for item in student.get("careerDocuments", {}).get("rounds", [])
        if event_in_range(item.get("date", ""), start_date, end_date)
    ]
    attendance_count = sum(1 for event in events if event.get("type") == "attendance")
    attendance_risk_count = sum(
        1
        for event in events
        if event.get("type") == "attendance" and event.get("scoreImpact") == "behavioral"
    )
    counseling_count = sum(1 for event in events if event.get("type") == "counseling")
    project_count = sum(1 for event in events if event.get("type") in {"project", "retro"})
    career_count = sum(1 for event in events if event.get("type") == "career")
    growth_delta = round(profile_avg - previous_average, 2) if previous_average is not None else 0.0
    sorted_scores = sorted(snapshot["scores"].items(), key=lambda item: item[1], reverse=True)
    strength_keys = [key for key, _ in sorted_scores[:2]]
    caution_keys = [key for key, value in sorted(snapshot["scores"].items(), key=lambda item: item[1])[:2] if value <= 2]

    return {
        "id": milestone["id"],
        "label": milestone["label"],
        "startDate": milestone["startDate"],
        "endDate": milestone["endDate"],
        "isEstimated": milestone["isEstimated"],
        "scores": snapshot["scores"],
        "profileAverage": profile_avg,
        "growthDelta": growth_delta,
        "strengthKeys": strength_keys,
        "cautionKeys": caution_keys,
        "confidence": snapshot["confidence"],
        "note": snapshot["note"],
        "eventCounts": {
            "attendance": attendance_count,
            "attendanceRisk": attendance_risk_count,
            "counseling": counseling_count,
            "project": project_count,
            "career": career_count,
        },
        "events": [
            {
                "id": event["id"],
                "date": event["date"],
                "title": event["title"],
                "summary": event["summary"],
                "severity": event["severity"],
                "type": event["type"],
            }
            for event in events[:5]
        ],
        "roles": [item.get("roleLabel", "") for item in role_history if item.get("roleLabel")],
        "careerRoundCount": len(career_rounds),
    }


def classify_student(student: dict[str, Any], growth_high_threshold: float) -> dict[str, Any]:
    milestone_summaries = student.get("milestones", [])
    first_average = milestone_summaries[0]["profileAverage"] if milestone_summaries else 0.0
    current_average = milestone_summaries[-1]["profileAverage"] if milestone_summaries else 0.0
    growth_delta = round(current_average - first_average, 2)
    positive_growth_steps = sum(1 for item in milestone_summaries if item.get("growthDelta", 0) > 0)
    caution_count = len([value for value in student["currentProfile"].values() if isinstance(value, int) and value <= 2])
    support_score = 0
    if student["stats"].get("attendanceRiskIssues", 0) >= 2:
        support_score += 1
    if student["stats"]["projectSubmissionRate"] < 70:
        support_score += 1
    if student["stats"]["currentStatus"] != "안정":
        support_score += 1
    if min(student["currentProfile"][key] for key in ["selfRegulation", "engagement", "collaboration", "resilience", "reflection", "careerAgency"]) <= 1:
        support_score += 1
    if caution_count >= 3:
        support_score += 1

    profile_index = round((current_average / 4) * 100) if current_average else 0
    growth_index = round(max(0, min(100, 50 + (growth_delta * 30))))
    support_index = round(
        max(0, min(100, support_score * 18 + student["stats"].get("attendanceRiskIssues", 0) * 10))
    )

    tags = []
    if current_average >= 3.15 and caution_count <= 1 and student["stats"]["currentStatus"] == "안정":
        tags.append("overall_strong")
    if growth_delta >= growth_high_threshold and current_average >= 3.0 and positive_growth_steps >= 2:
        tags.append("growth_high")
    if support_score >= 3:
        tags.append("support_priority")
    if student["currentProfile"]["collaboration"] >= 3 and (
        student["stats"]["leadershipRoleCount"] >= 1 or student["stats"]["repeatedPeerCount"] >= 2
    ):
        tags.append("collaboration_strength")
    if student["currentProfile"]["careerAgency"] >= 3 and student["stats"]["careerDocumentRounds"] >= 1:
        tags.append("career_progress")
    if student["stats"].get("attendanceRiskIssues", 0) >= 2 or student["currentProfile"]["engagement"] <= 2:
        tags.append("attendance_watch")
    if not tags:
        tags.append("steady_path")

    strength_keys = sorted(
        ["selfRegulation", "engagement", "collaboration", "resilience", "reflection", "careerAgency"],
        key=lambda key: student["currentProfile"][key],
        reverse=True,
    )[:2]
    caution_keys = [
        key
        for key in sorted(
            ["selfRegulation", "engagement", "collaboration", "resilience", "reflection", "careerAgency"],
            key=lambda key: student["currentProfile"][key],
        )[:2]
        if student["currentProfile"][key] <= 2
    ]

    primary_tag_order = [
        "support_priority",
        "overall_strong",
        "growth_high",
        "attendance_watch",
        "collaboration_strength",
        "career_progress",
        "steady_path",
    ]
    primary_tag = next((tag for tag in primary_tag_order if tag in tags), "steady_path")

    return {
        "profileAverage": current_average,
        "profileIndex": profile_index,
        "growthDelta": growth_delta,
        "growthIndex": growth_index,
        "supportIndex": support_index,
        "tags": tags,
        "primaryTag": primary_tag,
        "strengthKeys": strength_keys,
        "cautionKeys": caution_keys,
        "profileRankScore": round((current_average * 20) - (support_score * 4) + (student["currentProfile"]["reflection"] * 2), 2),
        "growthRankScore": round((growth_delta * 100) + (positive_growth_steps * 8), 2),
        "supportRankScore": round((support_score * 20) + (student["stats"].get("attendanceRiskIssues", 0) * 8), 2),
        "collaborationRankScore": round((student["currentProfile"]["collaboration"] * 20) + (student["stats"]["leadershipRoleCount"] * 6) + (student["stats"]["repeatedPeerCount"] * 4), 2),
        "careerRankScore": round((student["currentProfile"]["careerAgency"] * 20) + (student["stats"]["careerDocumentRounds"] * 12), 2),
    }


def build_dashboard_summary(students: list[dict[str, Any]], milestones: list[dict[str, Any]]) -> dict[str, Any]:
    stable_count = sum(1 for student in students if student["stats"]["currentStatus"] == "안정")
    caution_count = sum(1 for student in students if student["stats"]["currentStatus"] == "주의")
    warning_count = len(students) - stable_count - caution_count

    def pick_top(score_key: str, limit: int = 5) -> list[dict[str, Any]]:
        ordered = sorted(students, key=lambda item: item["derived"][score_key], reverse=True)
        return [
            {
                "id": student["id"],
                "name": student["name"],
                "score": student["derived"][score_key],
                "primaryTag": student["derived"]["primaryTag"],
                "strengthKeys": student["derived"]["strengthKeys"],
                "cautionKeys": student["derived"]["cautionKeys"],
                "profileIndex": student["derived"]["profileIndex"],
                "growthIndex": student["derived"]["growthIndex"],
            }
            for student in ordered[:limit]
        ]

    scatter_points = [
        {
            "id": student["id"],
            "name": student["name"],
            "x": student["derived"]["growthIndex"],
            "y": student["derived"]["profileIndex"],
            "primaryTag": student["derived"]["primaryTag"],
        }
        for student in students
    ]

    return {
        "studentCount": len(students),
        "stableCount": stable_count,
        "cautionCount": caution_count,
        "warningCount": warning_count,
        "studentsWithCareerDocuments": sum(1 for student in students if student["stats"]["careerDocumentRounds"] > 0),
        "milestoneCount": len(milestones),
        "topOverall": pick_top("profileRankScore"),
        "topGrowth": pick_top("growthRankScore"),
        "supportPriority": pick_top("supportRankScore"),
        "collaborationStrength": pick_top("collaborationRankScore"),
        "careerProgress": pick_top("careerRankScore"),
        "scatterPoints": scatter_points,
    }


def enrich_student_analysis(
    students: list[dict[str, Any]],
    curriculum: dict[str, Any],
    phase_dates: dict[str, list[date]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    phase_ranges = build_project_phase_ranges(curriculum, phase_dates)
    milestones = build_milestones(curriculum, phase_ranges, students)

    for student in students:
        milestone_summaries = []
        previous_average = None
        for milestone in milestones:
            summary = summarize_milestone(student, milestone, phase_dates, previous_average)
            milestone_summaries.append(summary)
            previous_average = summary.get("profileAverage", previous_average)
        student["milestones"] = milestone_summaries
        student["evaluationSnapshots"] = [
            {
                "snapshotId": f"{student['id']}-{item['id']}",
                "snapshotDate": item["endDate"],
                "snapshotType": item["label"],
                "scores": item["scores"],
                "confidence": item["confidence"],
                "note": item["note"],
            }
            for item in milestone_summaries
        ]
        latest_snapshot = student["evaluationSnapshots"][-1] if student["evaluationSnapshots"] else None
        if latest_snapshot:
            student["currentProfile"] = {
                **latest_snapshot["scores"],
                "confidence": latest_snapshot["confidence"],
                "note": latest_snapshot["note"],
                "updatedAt": latest_snapshot["snapshotDate"],
            }

    growth_high_threshold = max(
        0.75,
        percentile_value(
            [
                round(item["milestones"][-1]["profileAverage"] - item["milestones"][0]["profileAverage"], 2)
                for item in students
                if item.get("milestones")
            ],
            0.75,
        ),
    )

    for student in students:
        student["derived"] = classify_student(student, growth_high_threshold)

    dashboard = build_dashboard_summary(students, milestones)
    return phase_ranges, {
        "phaseRanges": phase_ranges,
        "milestones": milestones,
        "dashboard": dashboard,
    }


def build_payload() -> dict[str, Any]:
    students, students_by_name = build_students()
    curriculum, weeks = build_curriculum()
    course_start = parse_date(curriculum["startDate"])
    if course_start is None:
        raise RuntimeError("커리큘럼 시작일을 찾을 수 없습니다.")

    add_admission_data(students_by_name, course_start, weeks)
    phase_dates = add_project_data(students_by_name, weeks)
    add_project_team_data(students_by_name, weeks, phase_dates)
    add_attendance_data(students_by_name, weeks)
    add_counseling_data(students_by_name, weeks)
    add_career_document_data(students_by_name, weeks)
    build_evaluation_and_status(students, curriculum, weeks, phase_dates)
    phase_ranges, analysis = enrich_student_analysis(students, curriculum, phase_dates)

    students = sorted(students, key=lambda item: item["name"])
    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "curriculum": curriculum,
        "milestones": analysis["milestones"],
        "dashboard": analysis["dashboard"],
        "projectPhases": [
            {
                "phase": phase,
                "dates": [item.isoformat() for item in dates],
            }
            for phase, dates in phase_dates.items()
        ],
        "projectPhaseRanges": phase_ranges,
        "students": students,
    }
    return payload


def main() -> None:
    APP_DIR.mkdir(parents=True, exist_ok=True)
    payload = build_payload()
    json_text = json.dumps(payload, ensure_ascii=False, indent=2)
    OUTPUT_FILE.write_text(
        "window.STUDENT_TIMELINE_DATA = " + json_text + ";\n",
        encoding="utf-8",
    )
    print(f"Generated {OUTPUT_FILE}")
    print(f"Students: {len(payload['students'])}")
    print(f"Weeks: {len(payload['curriculum']['weeks'])}")


if __name__ == "__main__":
    main()
